"""
数据处理和导出模块，用于将结果导出为Excel
"""

import os
import pandas as pd
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime
import json
import io

import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import OUTPUT_DIR, DEFAULT_OUTPUT_FILENAME
from utils.text_formatter import format_problem_data
from utils.logger import get_logger
from utils.image_renderer import ImageRenderer

# 获取日志记录器
logger = get_logger("DataExporter")

class DataExporter:
    """数据导出类"""
    
    def __init__(self, output_dir: Optional[str] = None, filename: Optional[str] = None):
        """
        初始化数据导出器
        
        参数:
        - output_dir: 输出目录，如果为None则使用默认目录
        - filename: 输出文件名，如果为None则使用默认文件名
        """
        self.output_dir = output_dir or OUTPUT_DIR
        self.filename = filename or DEFAULT_OUTPUT_FILENAME
        
        # 确保输出目录存在
        os.makedirs(self.output_dir, exist_ok=True)
        
        # 初始化图像渲染器
        self.image_renderer = ImageRenderer(os.path.join(self.output_dir, 'rendered_images'))
    
    def export_to_excel(self, problems_with_solutions: List[Dict]) -> str:
        """
        将问题和解答导出为Excel
        
        参数:
        - problems_with_solutions: 包含问题和解答的字典列表
        
        返回:
        - 导出文件的路径
        """
        if not problems_with_solutions:
            logger.warning("No data to export")
            return ""
        
        try:
            # 创建DataFrame
            df = pd.DataFrame(problems_with_solutions)
            logger.debug(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
            
            # 重新排列列顺序
            columns = [
                "id", "question", "correctAnswer", "subject", 
                "grade", "type", "llmAnswer"
            ]
            
            # 确保所有列都存在
            for col in columns:
                if col not in df.columns:
                    logger.warning(f"Column {col} not found in DataFrame, adding empty column")
                    df[col] = ""
            
            # 选择并排序列
            df = df[columns]
            
            # 记录DataFrame的基本信息
            logger.debug(f"DataFrame columns: {df.columns.tolist()}")
            logger.debug(f"DataFrame shape: {df.shape}")
            
            # 检查correctAnswer列的内容
            empty_answers = df["correctAnswer"].isnull().sum() + (df["correctAnswer"] == "").sum()
            logger.info(f"Number of empty correctAnswer values: {empty_answers}")
            
            # 记录每个问题的correctAnswer
            for idx, row in df.iterrows():
                problem_id = row["id"]
                correct_answer = row["correctAnswer"]
                logger.debug(f"Problem {problem_id}: correctAnswer in DataFrame: {correct_answer}")
            
            # 生成带时间戳的文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename_parts = os.path.splitext(self.filename)
            timestamped_filename = f"{filename_parts[0]}_{timestamp}{filename_parts[1]}"
            
            # 导出路径
            export_path = os.path.join(self.output_dir, timestamped_filename)
            
            # 渲染LLM回答为图片
            logger.info("Rendering LLM answers to images")
            image_paths = []
            for idx, row in df.iterrows():
                problem_id = row["id"]
                llm_answer = row["llmAnswer"]
                
                # 渲染图片
                image_path = self.image_renderer.render_text_to_image(llm_answer, problem_id)
                image_paths.append(image_path)
            
            # 添加图片路径列
            df["llmAnswerImagePath"] = image_paths
            
            # 导出到Excel
            df.to_excel(export_path, index=False, engine='openpyxl')
            logger.debug(f"Exported DataFrame to {export_path}")
            
            # 使用openpyxl添加格式化和图片
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
            
            # 加载工作簿
            wb = load_workbook(export_path)
            ws = wb.active
            
            # 设置样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # 在Excel中添加一个新列用于图片
            ws.insert_cols(8)  # 在llmAnswer列之后插入一列
            ws.cell(row=1, column=8).value = "llmAnswerImage"  # 设置列标题
            ws.cell(row=1, column=8).font = header_font
            ws.cell(row=1, column=8).fill = header_fill
            ws.cell(row=1, column=8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=1, column=8).border = thin_border
            
            # 设置列宽
            for i, column in enumerate(columns, 1):
                col_letter = get_column_letter(i)
                if column == "question":
                    ws.column_dimensions[col_letter].width = 60
                elif column == "correctAnswer":
                    ws.column_dimensions[col_letter].width = 40  # 增加宽度以显示更多内容
                elif column == "llmAnswer":
                    ws.column_dimensions[col_letter].width = 60
                else:
                    ws.column_dimensions[col_letter].width = 15
            
            # 设置图片列宽度
            img_col_letter = get_column_letter(8)
            ws.column_dimensions[img_col_letter].width = 30
            
            # 设置行高，为图片预留空间
            for row_idx in range(2, len(df) + 2):
                ws.row_dimensions[row_idx].height = 200
            
            # 设置标题行样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # 设置所有单元格样式
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = thin_border
            
            # 添加图片到Excel
            for idx, image_path in enumerate(image_paths, 2):  # 从第2行开始（跳过标题行）
                if image_path and os.path.exists(image_path):
                    try:
                        # 创建图片对象并设置大小
                        img = XLImage(image_path)
                        
                        # 调整图片尺寸
                        img.width = 300  # 设置图片宽度
                        img.height = 180  # 设置图片高度
                        
                        # 添加图片到单元格
                        cell = ws.cell(row=idx, column=8)
                        cell_address = f"{img_col_letter}{idx}"
                        ws.add_image(img, cell_address)
                        
                        logger.debug(f"Added image at {cell_address}")
                    except Exception as e:
                        logger.error(f"Failed to add image to Excel: {str(e)}")
                        cell = ws.cell(row=idx, column=8)
                        cell.value = f"图片生成失败: {str(e)}"
            
            # 保存工作簿
            wb.save(export_path)
            
            logger.info(f"Data exported to {export_path}")
            return export_path
        
        except Exception as e:
            logger.error(f"Error exporting data to Excel: {str(e)}")
            logger.exception("Exception details:")
            return ""
    
    def format_data_for_export(self, problems: List[Dict], solutions: Dict[str, str]) -> List[Dict]:
        """
        格式化数据用于导出
        
        参数:
        - problems: 问题数据列表
        - solutions: 问题ID到解答的映射
        
        返回:
        - 格式化后的数据列表
        """
        formatted_data = []
        logger.info(f"Formatting {len(problems)} problems for export")
        
        for problem in problems:
            problem_id = problem.get("id", "")
            logger.debug(f"Formatting problem {problem_id} for export")
            logger.debug(f"Original problem data: {json.dumps(problem, ensure_ascii=False)}")
            
            # 创建新的字典，包含问题数据和解答
            item = problem.copy()
            
            # 添加大模型解答
            item["llmAnswer"] = solutions.get(problem_id, "")
            
            # 确保correctAnswer字段存在且不为空
            if not item.get("correctAnswer"):
                logger.warning(f"Problem {problem_id}: correctAnswer is empty or missing")
                item["correctAnswer"] = "答案不可用"
                logger.debug(f"Problem {problem_id}: Empty correctAnswer set to '答案不可用'")
            else:
                logger.debug(f"Problem {problem_id}: correctAnswer is {item['correctAnswer']}")
            
            # 应用文本格式化，使LaTeX公式更易读
            logger.debug(f"Applying text formatting for problem {problem_id}")
            item = format_problem_data(item)
            logger.debug(f"Formatted problem data: {json.dumps(item, ensure_ascii=False)}")
            
            formatted_data.append(item)
        
        logger.info(f"Formatted {len(formatted_data)} problems for export")
        return formatted_data 